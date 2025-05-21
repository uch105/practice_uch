from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl
import time

# create a new workbook
wb = openpyxl.Workbook()

# select the active worksheet
ws = wb.active

geckodriver_path = "/snap/bin/geckodriver"
driver_service = webdriver.FirefoxService(executable_path=geckodriver_path)
driver = webdriver.Firefox(service=driver_service)
#driver = webdriver.Firefox()
driver.get("http://dgdagov.info/index.php/registered-products/allopathic")
time.sleep(10)
'''
plist = driver.find_elements(By.CLASS_NAME,"Pharmaceutical")
blist = driver.find_elements(By.CLASS_NAME,"Brand")
glist = driver.find_elements(By.CLASS_NAME,"Generic")
slist = driver.find_elements(By.CLASS_NAME,"Strength")
dlist = driver.find_elements(By.CLASS_NAME,"Dosages")
pricelist = driver.find_elements(By.CLASS_NAME,"PRICE")
ulist = driver.find_elements(By.CLASS_NAME,"Use")
darlist = driver.find_elements(By.CLASS_NAME,"DAR")
next = driver.find_element(By.LINK_TEXT,"Next")
'''
count = 1

for i in range(1,1467):
    try:
        plist = driver.find_elements(By.CLASS_NAME,"Pharmaceutical")
        blist = driver.find_elements(By.CLASS_NAME,"Brand")
        glist = driver.find_elements(By.CLASS_NAME,"Generic")
        slist = driver.find_elements(By.CLASS_NAME,"Strength")
        dlist = driver.find_elements(By.CLASS_NAME,"Dosages")
        pricelist = driver.find_elements(By.CLASS_NAME,"PRICE")
        ulist = driver.find_elements(By.CLASS_NAME,"Use")
        darlist = driver.find_elements(By.CLASS_NAME,"DAR")
        for a in range(2,len(blist)):
            ws["A"+str(count)] = plist[a].text
            ws["B"+str(count)] = blist[a].text
            ws["C"+str(count)] = glist[a].text
            ws["D"+str(count)] = slist[a].text
            ws["E"+str(count)] = dlist[a].text
            ws["F"+str(count)] = pricelist[a].text
            ws["G"+str(count)] = ulist[a].text
            ws["H"+str(count)] = darlist[a].text
            count+=1
        print("Total "+str(count-1)+" Drugs printed!")
        next = driver.find_element(By.LINK_TEXT,"Next")
        next.click()
        time.sleep(7)
    except:
        plist = driver.find_elements(By.CLASS_NAME,"Pharmaceutical")
        blist = driver.find_elements(By.CLASS_NAME,"Brand")
        glist = driver.find_elements(By.CLASS_NAME,"Generic")
        slist = driver.find_elements(By.CLASS_NAME,"Strength")
        dlist = driver.find_elements(By.CLASS_NAME,"Dosages")
        pricelist = driver.find_elements(By.CLASS_NAME,"PRICE")
        ulist = driver.find_elements(By.CLASS_NAME,"Use")
        darlist = driver.find_elements(By.CLASS_NAME,"DAR")
        for a in range(2,len(blist)):
            ws["A"+str(count)] = plist[a].text
            ws["B"+str(count)] = blist[a].text
            ws["C"+str(count)] = glist[a].text
            ws["D"+str(count)] = slist[a].text
            ws["E"+str(count)] = dlist[a].text
            ws["F"+str(count)] = pricelist[a].text
            ws["G"+str(count)] = ulist[a].text
            ws["H"+str(count)] = darlist[a].text
            count+=1
        print("Total "+str(count-1)+" Drugs printed!")
        next = driver.find_element(By.LINK_TEXT,"Next")
        next.click()
        time.sleep(7)
wb.save('drugs.xlsx')
